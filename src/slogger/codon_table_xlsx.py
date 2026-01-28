from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
import openpyxl

@dataclass(frozen=True)
class CodonTable:
    aa_to_codons: Dict[str, List[Tuple[str, float]]]  # sorted desc by fraction, then codon

def load_codon_table_xlsx(path: str, sheet: Optional[str] = None) -> CodonTable:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in {path}. Available: {wb.sheetnames}")
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(x).strip().lower() if x is not None else "" for x in next(rows)]
    except StopIteration:
        raise ValueError("Empty codon table sheet.")

    def idx_of(*names: str) -> int:
        for n in names:
            if n in header:
                return header.index(n)
        return -1

    i_aa = idx_of("aa", "amino acid", "amino_acid")
    i_codon = idx_of("codon", "triplet")
    i_frac = idx_of("fraction", "freq", "frequency", "fraction_optimized", "usage", "codon_fraction")

    if i_aa < 0 or i_codon < 0:
        raise ValueError(f"Codon table must have AA and Codon columns; header={header}")

    tmp: Dict[str, List[Tuple[str, float]]] = {}
    for r in rows:
        aa = (str(r[i_aa]).strip().upper() if r[i_aa] is not None else "")
        codon = (str(r[i_codon]).strip().upper().replace("U","T") if r[i_codon] is not None else "")
        if not aa or not codon or len(codon) != 3:
            continue
        frac = 0.0
        if i_frac >= 0 and r[i_frac] is not None and str(r[i_frac]).strip() != "":
            try:
                frac = float(r[i_frac])
            except ValueError:
                frac = 0.0
        tmp.setdefault(aa, []).append((codon, frac))

    for aa in list(tmp.keys()):
        tmp[aa] = sorted(tmp[aa], key=lambda x: (-x[1], x[0]))

    return CodonTable(aa_to_codons=tmp)

def best_codons(table: CodonTable, aa: str) -> List[str]:
    aa = aa.upper()
    if aa not in table.aa_to_codons or not table.aa_to_codons[aa]:
        raise ValueError(f"No codons for AA={aa}")
    return [c for c,_ in table.aa_to_codons[aa]]
